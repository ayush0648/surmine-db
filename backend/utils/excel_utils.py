import openpyxl


def get_serial_number_and_update(sheet_path, details):
    """
    Fetch the next serial number from the Excel sheet and update it with details.

    Args:
        sheet_path (str): Path to the Excel sheet.
        details (dict): Record details to update.

    Returns:
        str: The serial number used.
    """
    wb = openpyxl.load_workbook(sheet_path)
    sheet = wb.active

    # Find the next available row
    next_row = sheet.max_row + 1
    serial_number = f"{next_row:04d}"  # Example: 0001

    # Write the details to the next row
    sheet.append([serial_number] + [details.get(col, "") for col in details.keys()])

    # Save the updated workbook
    wb.save(sheet_path)

    return serial_number


def fetch_records(sheet_path):
    """
    Fetch all records from an Excel sheet.

    Args:
        sheet_path (str): Path to the Excel sheet.

    Returns:
        list: List of records as dictionaries.
    """
    wb = openpyxl.load_workbook(sheet_path)
    sheet = wb.active

    records = []
    for row in sheet.iter_rows(values_only=True):
        records.append({"serial_number": row[0], "details": row[1:]})

    return records
